# 期末專案

import tkinter as tk
import tkinter.font as tkFont
import openpyxl
from PIL import ImageTk
import calendar
from tkinter import ttk
import tkinter.messagebox as tkmessage

root = tk.Tk()

class Page0:
    
    def _from_rgb(self, rgb):
        return "#%02x%02x%02x" % rgb 
    
    def __init__(self, master=None):
        self.root = master
        self.page0 = tk.Frame(self.root, width = 1000, height = 700)
        # self.page0.configure(bg = 'Dimgray')
        self.page0.configure(bg = self._from_rgb((68, 84, 106)))
        self.page0.master.title("開會小助手")
        self.page0.grid()
        self.createPage_0()


    def createPage_0(self):
        f1 = tkFont.Font(size = 45, family = "源泉圓體 B")
        f2 = tkFont.Font(size = 20, family = "源泉圓體 M")
        self.lblCaption = tk.Label(self.page0, text= '開 會 小 助 手', bg = self._from_rgb((68, 84, 106)), fg = 'white', font = f1, width = 30, height  = 2, anchor = 'center')
        self.btnStart = tk.Button(self.page0, text = '開始使用', bg = self._from_rgb((255, 217, 102)), font = f2, width = 10, height = 1, command = self.click_btnStart)
        
        self.lblCaption.place(relx = 0.5, rely = 0.5, anchor = 'center')
        self.btnStart.place(relx = 0.5, y = 450, anchor = 'center')
   
    def click_btnStart(self):
        self.page0.destroy()
        PageA()


class PageA:

    def _from_rgb(self, rgb):
        return "#%02x%02x%02x" % rgb 
    
    def __init__(self, master = None):
        self.root = master
        self.pageA = tk.Frame(self.root, width = 1000, height = 700)
        self.pageA.master.title("會議")
        self.pageA.grid()
        self.pageA.configure(bg = self._from_rgb((208, 224, 227)))
        self.createPage_A()
        self.count_meetings = 0
 
    def createPage_A(self):
        f1 = tkFont.Font(size = 30, family = "源泉圓體 B")
        f2 = tkFont.Font(size = 20, family = "源泉圓體 M")
        self.lblTitle_A = tk.Label(self.pageA, text = " 會議", height = 1 , width = 15, font = f1, bg = self._from_rgb((68, 84, 106)), fg = 'white', anchor = 'w')
        self.btnCreate_New = tk.Button(self.pageA, text = "創建新會議", height = 1, width = 10, font = f2, bg = self._from_rgb((255, 217, 102)), fg = 'black', command = self.click_btnCreate_New)
        
        self.lblTitle_A.place(x = 0, y = 50)
        self.btnCreate_New.place(x = 800, y = 550)
    
    def click_btnCreate_New(self):
        self.create_window()
        
    def create_window(self):
        f1 = tkFont.Font(size = 20, family = "源泉圓體 B")
        f2 = tkFont.Font(size = 15, family = "源泉圓體 M")
        
        self.window = tk.Toplevel()
        self.window.geometry('600x420')
        self.window.title('會議時間')
        self.text = tk.StringVar()
        self.text.set("測試")
        self.window.configure(bg = self._from_rgb((208, 224, 227)))
        self.lblTitle_B = tk.Label(self.window, text = " 創建新會議", height = 1 , width = 15, font = f1, bg = self._from_rgb((68, 84, 106)), fg = 'white', anchor = 'w')
        self.lblname = tk.Label(self.window, text = "會議名稱：", bg = self._from_rgb((208, 224, 227)), height = 1, width = 10, font = f2)
        self.lblchoose = tk.Label(self.window, text = "你已選擇：", bg = self._from_rgb((208, 224, 227)), height = 1, width = 10, font = f2)
        
        
        global meeting_name
        
        meeting_name = tk.StringVar()
        self.inputname = tk.Entry(self.window, textvariable = meeting_name, width = 30, font = f2)
        self.enydate = tk.Text(self.window,height = 7.4,width = 15,font = f2)
        width = root.winfo_reqwidth() + 50
        height = 100 #窗口大小
        x, y = (root.winfo_screenwidth()  - width )/2, (root.winfo_screenheight() - height)/2
       
        self.btnYes = tk.Button(self.window, text = "確認", height = 1, width = 5, bg = self._from_rgb((255, 217, 102)), font = f2, command = self.click_btnYes)
       
         
        self.lblTitle_B.place(x = 0, y = 25)
        self.lblname.place(x = 70, y = 75)
        self.lblchoose.place(x = 70, y = 105)
        self.enydate.place(x = 80, y = 140)
        self.inputname.place(x = 190, y = 75)
        self.btnYes.place(relx = 0.5, y = 380, anchor = 'center')
       
        datetime = calendar.datetime.datetime #日期和時間結合(從這邊複製)
        timedelta = calendar.datetime.timedelta  # 時間差
        
        class Calendar:
            def __init__(s, point = None, position = None):
                # point    窗口位置
                # position 窗口在點的位置 'ur'-右上, 'ul'-左上, 'll'-左下, 'lr'-右下
                fwday = calendar.SUNDAY 

                year = datetime.now().year  # 打開頁面時為當下年份
                month = datetime.now().month  # 打開頁面時為當下月份
                locale = None  # 地域設定
                sel_bg = '#ecffc4'  #點擊後框框色
                sel_fg = '#05640e'  # 點擊後字底色

                s._date = datetime(year, month, 1)  # 該月份第一天
                s._selection = None # 設置未選中的日期

                s.G_Frame = ttk.Frame(self.window)

                s._cal = s.__get_calendar(locale, fwday)   # 實例化適當的日曆類

                s.__setup_styles()       # 創建自定義樣式
                s.__place_widgets()      # pack/grid 小部件
                s.__config_calendar()    # 調整日曆列和安裝標記
                # 配置畫布和正確的绑定，以選擇日期。
                s.__setup_selection(sel_bg, sel_fg)
                # 存儲項ID，用於稍後插入。
                s._items = [s._calendar.insert('', 'end', values='') for _ in range(6)]

                # 在當前空日曆中插入日期
                s._update()

                s.G_Frame.place(x=290,y = 120)
                self.window.update_idletasks()  # 刷新頁面
               
                
                self.window.deiconify()  # 還原視窗 
                self.window.focus_set()  # 焦點設置在所需小部件上
                self.window.wait_window()  # 直到按確定
                
                
            def __get_calendar(s, locale, fwday):  # 日曆文字化
                if locale is None:
                    return calendar.TextCalendar(fwday)
                else:
                    return calendar.LocaleTextCalendar(fwday, locale)
                    
            def __setup_styles(s): # 自定義TTK風格
                style = ttk.Style(self.window)
                arrow_layout = lambda dir: ([('Button.focus', {'children': [('Button.%sarrow' % dir, None)]})])  # 返回參數性質
                style.layout('L.TButton', arrow_layout('left'))  #製作點選上個月的箭頭
                style.layout('R.TButton', arrow_layout('right'))  # 製作點選下個月的箭頭
                
            
            def __place_widgets(s):  # 標題框架及其小部件
                Input_judgment_num = self.window.register(s.Input_judgment)  # 需要将函数包装一下，必要的
                hframe = ttk.Frame(s.G_Frame)
                gframe = ttk.Frame(s.G_Frame)
                bframe = ttk.Frame(s.G_Frame)
                hframe.pack(in_=s.G_Frame, side='top',  pady=5, anchor='center')  # 日曆的上視窗
                gframe.pack(in_=s.G_Frame, fill=tk.X, pady=5)
                bframe.pack(in_=s.G_Frame, side='bottom', pady=5)  

                lbtn = ttk.Button(hframe, style='L.TButton', command=s._prev_month)  # 左箭頭
                lbtn.grid(in_=hframe, column=0, row=0, padx=12)
                rbtn = ttk.Button(hframe, style='R.TButton', command=s._next_month)  # 右箭頭
                rbtn.grid(in_=hframe, column=5, row=0, padx=12)
                
                s.CB_year = ttk.Combobox(hframe, width = 5, values = [str(year) for year in range(datetime.now().year, datetime.now().year+11,1)], validate = 'key', validatecommand = (Input_judgment_num, '%P'))  # 製作下拉選單
                s.CB_year.current(0)  # 下拉式索引起初在該年分
                s.CB_year.grid(in_=hframe, column=1, row=0)
                s.CB_year.bind("<<ComboboxSelected>>", s._update)
                tk.Label(hframe, text = '年', justify = 'left').grid(in_=hframe, column=2, row=0, padx=(0,5))  # 下拉式選單後面的單位(年)

                s.CB_month = ttk.Combobox(hframe, width = 3, values = ['%02d' % month for month in range(1,13)], state = 'readonly')  # 下拉式選單-月
                s.CB_month.current(datetime.now().month - 1)
                s.CB_month.grid(in_=hframe, column=3, row=0)
                s.CB_month.bind("<<ComboboxSelected>>", s._update)
                tk.Label(hframe, text = '月', justify = 'left').grid(in_=hframe, column=4, row=0)  # 下拉式選單後面的單位(年)
                # 日曆部件
                s._calendar = ttk.Treeview(gframe, show='', selectmode='none', height=7)
                s._calendar.pack(expand=1, fill='both', side='bottom', padx=5)
                ttk.Button(bframe, text = "加入", width = 6, command = lambda: s._exit(True)).grid(row = 0, column = 0, sticky = 'ns', padx = 20)  # 框架
                tk.Frame(s.G_Frame, bg = '#565656').place(x = 0, y = 0, relx = 0, rely = 0, relwidth = 1, relheigh = 2/200)
                tk.Frame(s.G_Frame, bg = '#565656').place(x = 0, y = 0, relx = 0, rely = 198/200, relwidth = 1, relheigh = 2/200)
                tk.Frame(s.G_Frame, bg = '#565656').place(x = 0, y = 0, relx = 0, rely = 0, relwidth = 2/200, relheigh = 1)
                tk.Frame(s.G_Frame, bg = '#565656').place(x = 0, y = 0, relx = 198/200, rely = 0, relwidth = 2/200, relheigh = 1)
                
            def __config_calendar(s):  # 設計日曆架構
                cols = ['日','一','二','三','四','五','六']  # 日曆上的星期幾
                s._calendar['columns'] = cols  # 設定日曆欄
                s._calendar.tag_configure('header', background='grey90')
                s._calendar.insert('', 'end', values=cols, tag='header')  # 調整其列寬
                font = tkFont.Font()
                maxwidth = max(font.measure(col) for col in cols)
                for col in cols:
                    s._calendar.column(col, width=maxwidth, minwidth=maxwidth,anchor='center')

            def __setup_selection(s, sel_bg, sel_fg):
                def __canvas_forget(evt):
                    canvas.place_forget()
                    s._selection = None

                s._font = tkFont.Font()
                s._canvas = canvas = tk.Canvas(s._calendar, background=sel_bg, borderwidth=0, highlightthickness=0)
                canvas.text = canvas.create_text(0, 0, fill=sel_fg, anchor='w')

                
                s._calendar.bind('<Button-1>', s._pressed)  #點出要的日期
          
                
            def _build_calendar(s):
                year, month = s._date.year, s._date.month  # update header text (Month, YEAR)
                header = s._cal.formatmonthname(year, month, 0)  # 更新日曆顯示的日期
                cal = s._cal.monthdayscalendar(year, month)
                for indx, item in enumerate(s._items):
                    week = cal[indx] if indx < len(cal) else []
                    fmt_week = [('%02d' % day) if day else '' for day in week]
                    s._calendar.item(item, values=fmt_week)
                    
            def _show_select(s, text, bbox):  # 秀出挑選的日子
                x, y, width, height = bbox

                textw = s._font.measure(text)
                canvas = s._canvas
                canvas.configure(width = width, height = height)
                canvas.coords(canvas.text, (width - textw)/2, height / 2 - 1)
                canvas.itemconfigure(canvas.text, text=text)
                canvas.place(in_=s._calendar, x=x, y=y)
                
            def _pressed(s, evt = None, item = None, column = None, widget = None):  #在日曆的某個地方點擊。
                
                if not item:
                    x, y, widget = evt.x, evt.y, evt.widget
                    item = widget.identify_row(y)
                    column = widget.identify_column(x)
                if not column or not item in s._items:  # 在工作日行中單擊或僅在列外單擊。
                    return
                item_values = widget.item(item)['values']  # 點選的日期該周
                if not len(item_values): # 該行是空的。
                    return
                text = item_values[int(column[1]) - 1] # text 為選擇日期
                if not text: # 日期為空
                    return
                bbox = widget.bbox(item, column)
                if not bbox: # 日曆尚未出現
                    self.window.after(20, lambda : s._pressed(item = item, column = column, widget = widget))
                    return
                text = '%02d' % text
                s._selection = (text, item, column)
                s._show_select(text, bbox)
            def _prev_month(s):
                s._canvas.place_forget()
                s._selection = None

                s._date = s._date - timedelta(days=1)
                s._date = datetime(s._date.year, s._date.month, 1)
                s.CB_year.set(s._date.year)
                s.CB_month.set(s._date.month)
                s._update()

            def _next_month(s):
                s._canvas.place_forget()
                s._selection = None

                year, month = s._date.year, s._date.month
                s._date = s._date + timedelta(
                    days=calendar.monthrange(year, month)[1] + 1)
                s._date = datetime(s._date.year, s._date.month, 1)
                s.CB_year.set(s._date.year)
                s.CB_month.set(s._date.month)
                s._update()

            def _update(s, event = None, key = None):
                if key and event.keysym != 'Return': return
                year = int(s.CB_year.get())
                month = int(s.CB_month.get())
                if year == 0 or year > 9999: return
                s._canvas.place_forget()
                s._date = datetime(year, month, 1)
                s._build_calendar() # 重建日曆

                if year == datetime.now().year and month == datetime.now().month:
                    day = datetime.now().day
                    for _item, day_list in enumerate(s._cal.monthdayscalendar(year, month)):
                        if day in day_list:
                            item = 'I00' + str(_item + 2)
                            column = '#' + str(day_list.index(day)+1)
                            self.window.after(100, lambda :s._pressed(item = item, column = column, widget = s._calendar))
                        
           
            def _exit(s, confirm = False,n = 0):
                
                global date_list
                date_list = []
                if not confirm:
                    s._selection = None
                year = s._date.year
                month = s._date.month
                self.enydate.insert(1.0,(str(year)+"/"+str(month)+"/"+str(int(s._selection[0]))+"\n"))
                if len(date_list) == 0:
                    date_list.append(self.enydate.get(1.0,"end"))
                else:
                    change_date = date_list[0].replace(" ","").split("\n")
                    for i in range(len(change_date)):
                        if len(change_date[i]) == 10:
                                if change_date[i][:11] == self.enydate.get(1.0,2.0):
                                    tkmessage.showerror(title="日期重複", message="此日期已選擇")
                                else:
                                    date_list.append(self.enydate.get(1.0,"end"))
                        
                    
                

            def _main_judge(s):
                try:
                    if self.window.focus_displayof() == None or 'toplevel' not in str(self.window.focus_displayof()): s._exit()
                    else: self.window.after(10, s._main_judge)
                except:
                    self.window.after(10, s._main_judge)
            def selection(s):
                if not s._selection: 
                    return None
                year = s._date.year
                month = s._date.month
                return str(datetime(year, month, int(s._selection[0])))[:10]

            def Input_judgment(s, content):
                if content.isdigit() or content == "":
                    return True
                else:
                    return False
            
            

        cal = Calendar()
        
        
    def click_btnYes(self):
        self.window.destroy()
        self.add_meetings()
        self.count_meetings += 1
        
        wb = openpyxl.load_workbook('C:/Users/sunny/Desktop/日期.xlsx')
        sheet = wb['工作表1']
        change_date = date_list[0].replace(" ","").split("\n")
        for i in range(len(change_date)-2):
            sheet.cell(row=1,column=i+2).value = change_date[i]
        wb.save('C:/Users/sunny/Desktop/日期.xlsx')
     
        
        

    def add_meetings(self):
        f1 = tkFont.Font(size = 30, family = "源泉圓體 B")
        f2 = tkFont.Font(size = 15, family = "源泉圓體 M")
        
        self.newMeeting = tk.Button(self.pageA, text = meeting_name.get(), relief = "solid", height = 2, width = 10, font = f1)
        self.newMeeting.place(x = 50 + 325*int(self.count_meetings%3), y = 150 + 150*int(self.count_meetings/3))
    
        data = openpyxl.load_workbook('C:/Users/sunny/Desktop/日期.xlsx')
        sheet = data.create_sheet(meeting_name.get(), 0)
        # 將工作表命名為會議名稱
       
        data.save('C:/Users/sunny/Desktop/日期.xlsx')


root.geometry("1000x700")
root.resizable(0, 0)
Page0(root)

root.mainloop()



