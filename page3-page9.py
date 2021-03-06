import tkinter as tk
import tkinter.font as tkfont
import tkinter.messagebox as tkmessage
import openpyxl

wb = openpyxl.load_workbook('/Users/jenny/Desktop/商管程專案/紀錄資料.xlsx')
absence = []
mission = []

root = tk.Tk()


class Page3:
    def __init__(self, master=None):
        self.root = master
        self.page3 = tk.Frame(self.root, width=1000, height=700)
        self.page3.master.title("Page3")
        self.page3.grid()

        f1 = tkfont.Font(size=12, family="Yu Gothic Medium")
        f2 = tkfont.Font(size=18, family="Yu Gothic Medium")

        self.btn_createtime = tk.Button(self.page3, text="新增你的時間", height=1, width=18, font=f2, command=self.click_btn_createtime)
        self.btn_times = tk.Button(self.page3, text="查看所有人的時間", height=1, width=18, font=f2, command=self.click_btn_times)
        self.btn_meetingrecord = tk.Button(self.page3, text="紀錄會議", height=1, width=18, font=f2, command=self.click_btn_meetingrecord)
        self.btn_back = tk.Button(self.page3, text="返回", height=1, font=f1, command=self.click_btn_back)

        self.btn_createtime.place(x=220, y=200)
        self.btn_times.place(x=505, y=200)
        self.btn_meetingrecord.place(x=360, y=275)
        self.btn_back.place(x=470, y=500)

    def click_btn_createtime(self):
        self.page3.destroy()
        Page4()

    def click_btn_times(self):
        self.page3.destroy()
        Page5()

    def click_btn_meetingrecord(self):
        self.page3.destroy()
        Page6()

    def click_btn_back(self):
        pass


class Page4:
    def __init__(self, master=None):
        self.root = master
        self.page4 = tk.Frame(self.root, width=1000, height=700)
        self.page4.master.title("Page4")
        self.page4.grid()

        f1 = tkfont.Font(size=12, family="Yu Gothic Medium")
        f2 = tkfont.Font(size=16, family="Yu Gothic Medium")

        global var_name, var_selectall

        self.lab_name = tk.Label(self.page4, text='姓名：', font=f2)
        var_name = tk.StringVar()
        self.int_name = tk.Entry(self.page4, textvariable=var_name, width=18, font=f2)
        self.btn_yes = tk.Button(self.page4, text='確定', height=1, font=f1, command=self.click_btn_yes)
        self.btn_back = tk.Button(self.page4, text='返回', height=1, font=f1, command=self.click_btn_back)
        var_selectall = tk.IntVar()
        self.btn_selectall = tk.Checkbutton(self.page4, onvalue=1, offvalue=0, variable=var_selectall, font=f1, command=self.click_selectall)
        self.lab_selectall = tk.Label(self.page4, text='全選', font=f1)

        self.lab_name.place(x=100, y=160)
        self.int_name.place(x=100, y=200, height=30)
        self.btn_back.place(x=100, y=250)
        self.btn_yes.place(x=475, y=620)
        self.btn_selectall.place(x=855, y=49)
        self.lab_selectall.place(x=878, y=50)

        global chk_btns
        chk_btns = []

        for i in range(8):
            if i != 0:
                chk_btns.append([])
            for j in range(17):
                if i == 0:
                    if j == 0:
                        tk.Label(self.page4, relief='solid', borderwidth=1, width=10, height=2).place(x=480, y=80)
                    else:
                        tk.Label(self.page4, text=str(6 + j) + ':00-' + str(7 + j) + ':00', relief='solid', borderwidth=1, width=10, height=2).place(x=480, y=80 + 30 * j)
                else:
                    tk.Label(self.page4, relief='solid', borderwidth=1, width=7, height=2).place(x=501 + 52 * i, y=80 + 30 * j)
                    if j != 0:
                        var_i = tk.IntVar()
                        tk.Checkbutton(self.page4, onvalue=1, offvalue=0, variable=var_i).place(x=518 + 52 * i, y=85 + 30 * j)
                        chk_btns[i - 1].append(var_i)

    def click_btn_yes(self):
        if var_name.get() == "":
            tkmessage.showerror(title="請輸入姓名", message="請輸入姓名！")
        else:
            global wb
            sheet = wb['時間統計']

            member = str(sheet.cell(row=2, column=10).value)
            if member == 'None':
                list_member = [var_name.get()]
                for i in range(7):
                    for j in range(16):
                        if chk_btns[i][j].get() == 1:
                            available = str(sheet.cell(row=j + 2, column=i + 2).value)
                            if available == 'None':
                                list_available = [var_name.get()]
                            else:
                                list_available = available.split(',')
                                list_available.append(var_name.get())
                            sheet.cell(row=j + 2, column=i + 2).value = ",".join(list_available)
            else:
                list_member = member.split(',')
                if var_name.get() not in list_member:
                    list_member.append(var_name.get())
                    for i in range(7):
                        for j in range(16):
                            if chk_btns[i][j].get() == 1:
                                available = str(sheet.cell(row=j + 2, column=i + 2).value)
                                if available == 'None':
                                    list_available = [var_name.get()]
                                else:
                                    list_available = available.split(',')
                                    list_available.append(var_name.get())
                                sheet.cell(row=j + 2, column=i + 2).value = ",".join(list_available)
                else:
                    for i in range(7):
                        for j in range(16):
                            available = str(sheet.cell(row=j + 2, column=i + 2).value)
                            if chk_btns[i][j].get() == 1:
                                if available == 'None':
                                    list_available = [var_name.get()]
                                else:
                                    list_available = available.split(',')
                                    if var_name.get() not in available:
                                        list_available.append(var_name.get())
                                sheet.cell(row=j + 2, column=i + 2).value = ",".join(list_available)
                            else:
                                if var_name.get() in available:
                                    available = available.replace(var_name.get(), "").replace(",,", ",").strip(",")
                                    sheet.cell(row=j + 2, column=i + 2).value = available

            sheet.cell(row=2, column=10).value = ",".join(list_member)
            wb.save('/Users/jenny/Desktop/商管程專案/紀錄資料.xlsx')

            self.page4.destroy()
            Page5()

    def click_btn_back(self):
        self.page4.destroy()
        Page3()

    def click_selectall(self):
        if var_selectall.get() == 1:
            for i in range(7):
                for j in range(16):
                    chk_btns[i][j].set(1)
        else:
            for i in range(7):
                for j in range(16):
                    chk_btns[i][j].set(0)


class Page5:
    def __init__(self, master=None):
        self.root = master
        self.page5 = tk.Frame(self.root, width=1000, height=700)
        self.page5.master.title("Page5")
        self.page5.grid()

        f1 = tkfont.Font(size=12, family="Yu Gothic Medium")

        self.btn_yes = tk.Button(self.page5, text='確定', height=1, font=f1, command=self.click_btn_yes)
        self.lab_allmembers = tk.Label(self.page5, text='all members', font=f1)
        self.lab_available = tk.Label(self.page5, text='available', font=f1)
        self.lab_unavailable = tk.Label(self.page5, text='unavailable', font=f1)
        self.btn_yes.place(x=475, y=620)
        self.lab_allmembers.place(x=100, y=74)
        self.lab_available.place(x=319, y=256)
        self.lab_unavailable.place(x=103, y=256)

        var_available = tk.StringVar()
        self.lst_available = tk.Listbox(self.page5, listvariable=var_available, font=f1, width=15, height=14)
        self.lst_available.place(x=285, y=282)

        var_unavailable = tk.StringVar()
        self.lst_unavailable = tk.Listbox(self.page5, listvariable=var_unavailable, font=f1, width=15, height=14)
        self.lst_unavailable.place(x=80, y=282)

        var_allmembers = tk.StringVar()
        self.lst_allmembers = tk.Listbox(self.page5, listvariable=var_allmembers, font=f1, width=15, height=6)
        self.lst_allmembers.place(x=80, y=100)

        global wb
        sheet = wb['時間統計']
        all_members = str(sheet.cell(row=2, column=10).value).split(',')
        for member in all_members:
            self.lst_allmembers.insert("end", member)

        self.lst_color = tk.Listbox(self.page5, width=15, height=6, font=f1)
        self.lst_color.place(x=285, y=100)

        color = ["blue", "red", "green"]

        for i in range(3):
            self.lst_color.insert("end", i)
            self.lst_color.itemconfig("end", bg=color[i])

        global btn_list
        btn_list = []
        num = []

        for i in range(8):
            if i != 0:
                btn_list.append([])
                num.append([])
            for j in range(17):
                if i == 0:
                    if j == 0:
                        tk.Label(self.page5, relief='solid', borderwidth=1, width=10, height=2).place(x=480, y=80)
                    else:
                        tk.Label(self.page5, text=str(6 + j) + ':00-' + str(7 + j) + ':00', relief='solid', borderwidth=1, width=10, height=2).place(x=480, y=80 + 30 * j)
                else:
                    tk.Label(self.page5, relief='solid', borderwidth=1, width=7, height=2).place(x=501 + 52 * i, y=80 + 30 * j)
                    if j != 0:
                        self.btn = tk.Button(self.page5, background="#AAFFEE", width=6, height=1, command=lambda a=i, b=j: self.click_btn(a, b))
                        self.btn.place(x=502 + 52 * i, y=83 + 30 * j)
                        btn_list[i - 1].append(self.btn)

    def click_btn_yes(self):
        self.page5.destroy()
        Page3()

    def click_btn(self, a, b):
        self.lst_available.delete(0, "end")
        self.lst_unavailable.delete(0, "end")

        global wb
        sheet = wb['時間統計']
        all_members = str(sheet.cell(row=2, column=10).value).split(',')
        available_member = str(sheet.cell(row=b + 1, column=a + 1).value).split(',')

        for member in available_member:
            self.lst_available.insert("end", member)

        for member in all_members:
            if member not in available_member:
                self.lst_unavailable.insert("end", member)


class Page6:
    def __init__(self, master=None):
        self.root = master
        self.page6 = tk.Frame(self.root, width=1000, height=700)
        self.page6.master.title("Page6")
        self.page6.grid()

        f1 = tkfont.Font(size=12, family="Yu Gothic Medium")
        f2 = tkfont.Font(size=16, family="Yu Gothic Medium")

        self.btn6_1 = tk.Button(self.page6, text="出缺勤", height=1, width=18, font=f2, command=self.click_btn6_1)
        self.btn6_2 = tk.Button(self.page6, text="會議記錄", height=1, width=18, font=f2, command=self.click_btn6_2)
        self.btn6_3 = tk.Button(self.page6, text="結束會議", height=1, width=18, font=f2, command=self.click_btn6_3)
        self.btn6_4 = tk.Button(self.page6, text="返回", font=f1, command=self.click_btn6_4)

        self.btn6_1.place(x=380, y=200)
        self.btn6_2.place(x=380, y=275)
        self.btn6_3.place(x=380, y=350)
        self.btn6_4.place(x=470, y=500)

    def click_btn6_1(self):
        self.page6.destroy()
        Page7()

    def click_btn6_2(self):
        self.page6.destroy()
        Page8()

    def click_btn6_3(self):
        self.page6.destroy()
        Page9()

    def click_btn6_4(self):
        self.page6.destroy()
        Page3()


class Page7:
    def __init__(self, master=None):
        self.root = master
        self.page7 = tk.Frame(self.root, width=1000, height=700)
        self.page7.master.title("Page8")
        self.page7.grid()

        f1 = tkfont.Font(size=12, family="Yu Gothic Medium")
        f2 = tkfont.Font(size=16, family="Yu Gothic Medium")

        self.btn7_1 = tk.Button(self.page7, text="確定", font=f1, command=self.click_btn7_1)
        self.btn7_1.place(x=475, y=620)

        global wb
        times = wb['時間統計']
        try:
            new_sheet = wb['出缺勤']
        except KeyError:
            new_sheet = wb.create_sheet('出缺勤')

        global absence, mission, member_list

        member_list = str(times.cell(row=2, column=10).value).split(',')
        for i in range(len(member_list)):
            new_sheet.cell(row=i + 1, column=1).value = member_list[i]
            tk.Label(self.page7, text=member_list[i], font=f2).place(x=100, y=50 + 35 * i)
            tk.Label(self.page7, text='是否完成指派任務？', font=f2).place(x=500, y=50 + 35 * i)

            if len(absence) != len(member_list):
                var_absence = tk.IntVar()
                absence.append(var_absence)
            else:
                var_absence = absence[i]
            tk.Radiobutton(self.page7, variable=var_absence, text="準時", value=1, font=f2).place(x=175, y=50 + 35 * i)
            tk.Radiobutton(self.page7, variable=var_absence, text="遲到", value=2, font=f2).place(x=275, y=50 + 35 * i)
            tk.Radiobutton(self.page7, variable=var_absence, text="未出席", value=3, font=f2).place(x=375, y=50 + 35 * i)

            if len(mission) != len(member_list):
                var_mission = tk.IntVar()
                mission.append(var_mission)
            else:
                var_mission = mission[i]
            tk.Radiobutton(self.page7, variable=var_mission, text="是", value=1, font=f2).place(x=675, y=50 + 35 * i)
            tk.Radiobutton(self.page7, variable=var_mission, text="否", value=2, font=f2).place(x=745, y=50 + 35 * i)
            tk.Radiobutton(self.page7, variable=var_mission, text="無任務", value=3, font=f2).place(x=815, y=50 + 35 * i)

        wb.save('/Users/jenny/Desktop/商管程專案/紀錄資料.xlsx')

    def click_btn7_1(self):
        global wb
        sheet = wb['出缺勤']

        for i in range(len(member_list)):
            if absence[i].get() == 1:
                sheet.cell(row=i + 1, column=2).value = "準時"
            if absence[i].get() == 2:
                sheet.cell(row=i + 1, column=2).value = "遲到"
            if absence[i].get() == 3:
                sheet.cell(row=i + 1, column=2).value = "未出席"
            if mission[i].get() == 1:
                sheet.cell(row=i + 1, column=3).value = "完成任務"
            if mission[i].get() == 2:
                sheet.cell(row=i + 1, column=3).value = "未完成任務"
            if mission[i].get() == 3:
                sheet.cell(row=i + 1, column=3).value = "無任務"

        wb.save('/Users/jenny/Desktop/商管程專案/紀錄資料.xlsx')

        self.page7.destroy()
        Page6()


class Page8:
    def __init__(self, master=None):
        self.root = master
        self.page8 = tk.Frame(self.root, width=1000, height=700)
        self.page8.master.title("Page8")
        self.page8.grid()

        f1 = tkfont.Font(size=12, family="Yu Gothic Medium")
        f2 = tkfont.Font(size=16, family="Yu Gothic Medium")

        self.record = tk.Text(self.page8, height=23, width=90, font=f1)
        self.lbl8_1 = tk.Label(self.page8, text="會議記錄", font=f2)
        self.btn8 = tk.Button(self.page8, text='確認', font=f1, command=self.click_btn8)
        self.lbl8_2 = tk.Label(self.page8, text="會議時間：", font=f1)

        global var_times
        var_times = tk.StringVar()
        self.entry8 = tk.Entry(self.page8, textvariable=var_times, width=18, font=f1)

        global wb, ws
        try:
            ws = wb['Meeting record']
        except KeyError:
            ws = wb.create_sheet('Meeting record')
        if str(ws.cell(row=1, column=1).value) != 'None':
            self.record.insert("1.0", str(ws.cell(row=1, column=1).value))

        self.lbl8_1.place(x=450, y=40)
        self.record.place(x=95, y=115)
        self.btn8.place(x=475, y=620)
        self.entry8.place(x=743, y=80)
        self.lbl8_2.place(x=655, y=80)

    def click_btn8(self):
        ws.cell(row=1, column=1).value = self.record.get("1.0", "end")
        ws.cell(row=1, column=2).value = var_times.get()
        wb.save('/Users/jenny/Desktop/商管程專案/紀錄資料.xlsx')
        self.page8.destroy()
        Page6()

class Page9:
    def __init__(self, master=None):
        self.root = master
        self.page9 = tk.Frame(self.root, width=1000, height=700)
        self.page9.master.title("Page9")
        self.page9.grid()

        f1 = tkfont.Font(size=12, family="Yu Gothic Medium")
        f2 = tkfont.Font(size=16, family="Yu Gothic Medium")

        global wb
        # ws_1 = wb[''] 讀取會議名稱
        ws_2 = wb['Meeting record']
        ws_3 = wb['出缺勤']

        self.lbl9_1 = tk.Label(self.page9, text="會議名稱：", font=f2)
        # self.lbl9_2 = tk.Label(self.page9, text=)  會議名稱
        self.lbl9_3 = tk.Label(self.page9, text="會議時間：", font=f2)
        self.lbl9_4 = tk.Label(self.page9, text=str(ws_2.cell(row = 1, column = 2).value), font=f2)
        self.lbl9_5 = tk.Label(self.page9, text="會議記錄：", font=f2)
        self.lbl9_6 = tk.Label(self.page9, text=str(ws_2.cell(row = 1, column = 1).value), font=f2)
        self.lbl9_7 = tk.Label(self.page9, text="成員名單：", font=f2)
        self.lbl9_9 = tk.Label(self.page9, text="出缺勤：", font=f2)
        self.lbl9_11 = tk.Label(self.page9, text="是否完成指派任務？", font=f2)
        # self.btn9 = tk.Button(self.page9, text="返回", command=self.click_btn9)  返回至第二頁（創建新會議）的按鈕

        self.lbl9_1.place(x = 100, y = 50)
        # self.lbl9_2.place()
        self.lbl9_3.place(x = 100, y = 100)
        self.lbl9_4.place(x = 180, y = 100)
        self.lbl9_5.place(x = 100, y = 150)
        self.lbl9_6.place(x = 180, y = 150)
        self.lbl9_7.place(x=100, y=300)
        self.lbl9_9.place(x=100, y=350)
        self.lbl9_11.place(x=100, y=400)
        # self.btn9.place(x=475, y=620)

    # def click_btn9_1(self):
    #     self.page9.destroy()
    #     Page2()

        times = wb['時間統計']
        member_list = str(times.cell(row=2, column=10).value).split(',')
        for i in range(len(member_list)):
            self.lbl9_8 = tk.Label(self.page9, text=str(ws_3.cell(row=i + 1, column=1).value), font=f2).place(x=280 + 70*i, y=300)
            self.lbl9_10 = tk.Label(self.page9, text=str(ws_3.cell(row=i + 1, column=2).value), font=f2).place(x=280 + 70*i, y=350)
            self.lbl9_12 = tk.Label(self.page9, text=str(ws_3.cell(row=i + 1, column=3).value), font=f2).place(x=280 + 70*i, y=400)




root.geometry("1000x700")
root.resizable(0, 0)
Page3(root)
root.mainloop()