import tkinter as tk
import tkinter.font as tkfont
import tkinter.messagebox as tkmessage
import openpyxl
import calendar
from tkinter import ttk

colors = ['red4', 'firebrick4', 'firebrick3', 'red2', 'red', 'firebrick1', 'OrangeRed2', 'tomato2', 'tomato',
          'chocolate1', 'dark orange', 'orange', 'goldenrod1', 'gold', 'yellow', 'DarkOliveGreen1', 'OliveDrab1',
          'green yellow', 'lawn green', 'chartreuse2', 'lime green', 'green3', 'SpringGreen3', 'SeaGreen3',
          'medium sea green', 'springGreen4', 'sea green', 'forest green', 'green4', 'dark green']


absence = []
mission = []


root = tk.Tk()


class Page0:

    def _from_rgb(self, rgb):
        return "#%02x%02x%02x" % rgb

    def __init__(self, master=None):
        self.root = master
        self.page0 = tk.Frame(self.root, width=1000, height=700)
        self.page0.configure(bg=self._from_rgb((68, 84, 106)))
        self.page0.master.title("開會小助手")
        self.page0.grid()

        f1 = tkfont.Font(size=45, family="源泉圓體 B")
        f2 = tkfont.Font(size=20, family="源泉圓體 M")
        self.lblCaption = tk.Label(self.page0, text='開 會 小 助 手', bg=self._from_rgb((68, 84, 106)), fg='white',
                                   font=f1, width=30, height=2, anchor='center')
        self.btnStart = tk.Button(self.page0, text='開始使用', bg=self._from_rgb((255, 217, 102)), font=f2, width=10,
                                  height=1, command=self.click_btnStart)

        self.lblCaption.place(relx=0.5, rely=0.5, anchor='center')
        self.btnStart.place(relx=0.5, y=450, anchor='center')

    def click_btnStart(self):
        self.page0.destroy()
        Page1()


class Page1:

    def _from_rgb(self, rgb):
        return "#%02x%02x%02x" % rgb

    def __init__(self, master=None):
        self.root = master
        self.page1 = tk.Frame(self.root, width=1000, height=700, bg=self._from_rgb((208, 224, 227)))
        self.page1.master.title("會議")
        self.page1.grid()
        self.count_meetings = 0

        f1 = tkfont.Font(size=30, family="源泉圓體 B")
        f2 = tkfont.Font(size=20, family="源泉圓體 M")

        self.lblTitle_A = tk.Label(self.page1, text=" 會議", height=1, width=15, font=f1,
                                   bg=self._from_rgb((68, 84, 106)), fg='white', anchor='w')
        self.btnCreate_New = tk.Button(self.page1, text="創建新會議", height=1, width=10, font=f2,
                                       bg=self._from_rgb((255, 217, 102)), fg='black', command=self.click_btnCreate_New)

        self.lblTitle_A.place(x=0, y=50)
        self.btnCreate_New.place(x=800, y=550)

        global names, sheet_names
        names = openpyxl.load_workbook('會議名稱.xlsx')
        sheet_names = names['會議']

        global meeting_names, finish_meeting
        meeting_names = []
        finish_meeting = []

        try:
            for cell in list(sheet_names.columns)[0]:
                meeting_names.append(cell.value)
            for cell in list(sheet_names.columns)[1]:
                finish_meeting.append(cell.value)
        except IndexError:
            pass

        for i in range(len(meeting_names)):
            self.btn_names = tk.Button(self.page1, text=meeting_names[i], height=2, width=10, relief='solid', font=f1,
                                       bg='white', command=lambda a=i: self.click_btn_meetings(a))
            self.btn_names.place(x=44 + 325 * (i % 3), y=150 + 150 * (i // 3))

    def click_btnCreate_New(self):
        self.create_window()

    def create_window(self):
        f1 = tkfont.Font(size=20, family="源泉圓體 B")
        f2 = tkfont.Font(size=15, family="源泉圓體 M")

        self.window = tk.Toplevel()
        self.window.geometry('600x420')
        self.window.title('會議時間')
        self.text = tk.StringVar()
        self.text.set("測試")
        self.window.configure(bg=self._from_rgb((208, 224, 227)))
        self.lblTitle_B = tk.Label(self.window, text=" 創建新會議", height=1, width=15, font=f1,
                                   bg=self._from_rgb((68, 84, 106)), fg='white', anchor='w')
        self.lblname = tk.Label(self.window, text="會議名稱：", bg=self._from_rgb((208, 224, 227)), height=1, width=10,
                                font=f2)
        self.lblchoose = tk.Label(self.window, text="你已選擇：", bg=self._from_rgb((208, 224, 227)), height=1, width=10,
                                  font=f2)

        global meeting_name

        meeting_name = tk.StringVar()
        self.inputname = tk.Entry(self.window, textvariable=meeting_name, width=30, font=f2)
        self.enydate = tk.Text(self.window, height=7.4, width=15, font=f2)
        width = root.winfo_reqwidth() + 50
        height = 100  # 窗口大小
        x, y = (root.winfo_screenwidth() - width) / 2, (root.winfo_screenheight() - height) / 2

        self.btnYes = tk.Button(self.window, text="確認", height=1, width=5, bg=self._from_rgb((255, 217, 102)), font=f2,
                                command=self.click_btnYes)

        self.lblTitle_B.place(x=0, y=25)
        self.lblname.place(x=70, y=75)
        self.lblchoose.place(x=70, y=105)
        self.enydate.place(x=80, y=140)
        self.inputname.place(x=190, y=75)
        self.btnYes.place(relx=0.5, y=380, anchor='center')

        datetime = calendar.datetime.datetime  # 日期和時間結合(從這邊複製)
        timedelta = calendar.datetime.timedelta  # 時間差

        class Calendar:
            def __init__(s, point=None, position=None):
                # point    窗口位置
                # position 窗口在點的位置 'ur'-右上, 'ul'-左上, 'll'-左下, 'lr'-右下
                fwday = calendar.SUNDAY

                year = datetime.now().year  # 打開頁面時為當下年份
                month = datetime.now().month  # 打開頁面時為當下月份
                locale = None  # 地域設定
                sel_bg = '#ecffc4'  # 點擊後框框色
                sel_fg = '#05640e'  # 點擊後字底色

                s._date = datetime(year, month, 1)  # 該月份第一天
                s._selection = None  # 設置未選中的日期

                s.G_Frame = ttk.Frame(self.window)

                s._cal = s.__get_calendar(locale, fwday)  # 實例化適當的日曆類

                s.__setup_styles()  # 創建自定義樣式
                s.__place_widgets()  # pack/grid 小部件
                s.__config_calendar()  # 調整日曆列和安裝標記
                # 配置畫布和正確的绑定，以選擇日期。
                s.__setup_selection(sel_bg, sel_fg)
                # 存儲項ID，用於稍後插入。
                s._items = [s._calendar.insert('', 'end', values='') for _ in range(6)]

                # 在當前空日曆中插入日期
                s._update()

                s.G_Frame.place(x=290, y=120)
                self.window.update_idletasks()  # 刷新頁面

                self.window.deiconify()  # 還原視窗
                self.window.focus_set()  # 焦點設置在所需小部件上
                self.window.wait_window()  # 直到按確定

            def __get_calendar(s, locale, fwday):  # 日曆文字化
                if locale is None:
                    return calendar.TextCalendar(fwday)
                else:
                    return calendar.LocaleTextCalendar(fwday, locale)

            def __setup_styles(s):  # 自定義TTK風格
                style = ttk.Style(self.window)
                arrow_layout = lambda dir: (
                    [('Button.focus', {'children': [('Button.%sarrow' % dir, None)]})])  # 返回參數性質
                style.layout('L.TButton', arrow_layout('left'))  # 製作點選上個月的箭頭
                style.layout('R.TButton', arrow_layout('right'))  # 製作點選下個月的箭頭

            def __place_widgets(s):  # 標題框架及其小部件
                Input_judgment_num = self.window.register(s.Input_judgment)  # 需要将函数包装一下，必要的
                hframe = ttk.Frame(s.G_Frame)
                gframe = ttk.Frame(s.G_Frame)
                bframe = ttk.Frame(s.G_Frame)
                hframe.pack(in_=s.G_Frame, side='top', pady=5, anchor='center')  # 日曆的上視窗
                gframe.pack(in_=s.G_Frame, fill=tk.X, pady=5)
                bframe.pack(in_=s.G_Frame, side='bottom', pady=5)

                lbtn = ttk.Button(hframe, style='L.TButton', command=s._prev_month)  # 左箭頭
                lbtn.grid(in_=hframe, column=0, row=0, padx=12)
                rbtn = ttk.Button(hframe, style='R.TButton', command=s._next_month)  # 右箭頭
                rbtn.grid(in_=hframe, column=5, row=0, padx=12)

                s.CB_year = ttk.Combobox(hframe, width=5, values=[str(year) for year in
                                                                  range(datetime.now().year, datetime.now().year + 11,
                                                                        1)], validate='key',
                                         validatecommand=(Input_judgment_num, '%P'))  # 製作下拉選單
                s.CB_year.current(0)  # 下拉式索引起初在該年分
                s.CB_year.grid(in_=hframe, column=1, row=0)
                s.CB_year.bind("<<ComboboxSelected>>", s._update)
                tk.Label(hframe, text='年', justify='left').grid(in_=hframe, column=2, row=0,
                                                                padx=(0, 5))  # 下拉式選單後面的單位(年)

                s.CB_month = ttk.Combobox(hframe, width=3, values=['%02d' % month for month in range(1, 13)],
                                          state='readonly')  # 下拉式選單-月
                s.CB_month.current(datetime.now().month - 1)
                s.CB_month.grid(in_=hframe, column=3, row=0)
                s.CB_month.bind("<<ComboboxSelected>>", s._update)
                tk.Label(hframe, text='月', justify='left').grid(in_=hframe, column=4, row=0)  # 下拉式選單後面的單位(年)
                # 日曆部件
                s._calendar = ttk.Treeview(gframe, show='', selectmode='none', height=7)
                s._calendar.pack(expand=1, fill='both', side='bottom', padx=5)
                ttk.Button(bframe, text="加入", width=6, command=lambda: s._exit(True)).grid(row=0, column=0, sticky='ns',
                                                                                           padx=20)
                tk.Frame(s.G_Frame, bg='#565656').place(x=0, y=0, relx=0, rely=0, relwidth=1, relheigh=2 / 200)
                tk.Frame(s.G_Frame, bg='#565656').place(x=0, y=0, relx=0, rely=198 / 200, relwidth=1, relheigh=2 / 200)
                tk.Frame(s.G_Frame, bg='#565656').place(x=0, y=0, relx=0, rely=0, relwidth=2 / 200, relheigh=1)
                tk.Frame(s.G_Frame, bg='#565656').place(x=0, y=0, relx=198 / 200, rely=0, relwidth=2 / 200, relheigh=1)

            def __config_calendar(s):  # 設計日曆架構
                cols = ['日', '一', '二', '三', '四', '五', '六']  # 日曆上的星期幾
                s._calendar['columns'] = cols  # 設定日曆欄
                s._calendar.tag_configure('header', background='grey90')
                s._calendar.insert('', 'end', values=cols, tag='header')  # 調整其列寬
                font = tkfont.Font()
                maxwidth = max(font.measure(col) for col in cols)
                for col in cols:
                    s._calendar.column(col, width=maxwidth, minwidth=maxwidth, anchor='center')

            def __setup_selection(s, sel_bg, sel_fg):
                def __canvas_forget(evt):
                    canvas.place_forget()
                    s._selection = None

                s._font = tkfont.Font()
                s._canvas = canvas = tk.Canvas(s._calendar, background=sel_bg, borderwidth=0, highlightthickness=0)
                canvas.text = canvas.create_text(0, 0, fill=sel_fg, anchor='w')

                s._calendar.bind('<Button-1>', s._pressed)  # 點出要的日期

            def _build_calendar(s):
                year, month = s._date.year, s._date.month  # update header text (Month, YEAR)
                header = s._cal.formatmonthname(year, month, 0)  # 更新日曆顯示的日期
                cal = s._cal.monthdayscalendar(year, month)
                for indx, item in enumerate(s._items):
                    week = cal[indx] if indx < len(cal) else []
                    fmt_week = [('%02d' % day) if day else '' for day in week]
                    s._calendar.item(item, values=fmt_week)

            def _show_select(s, text, bbox):  # 秀出挑選的日子
                x, y, width, height = bbox

                textw = s._font.measure(text)
                canvas = s._canvas
                canvas.configure(width=width, height=height)
                canvas.coords(canvas.text, (width - textw) / 2, height / 2 - 1)
                canvas.itemconfigure(canvas.text, text=text)
                canvas.place(in_=s._calendar, x=x, y=y)

            def _pressed(s, evt=None, item=None, column=None, widget=None):  # 在日曆的某個地方點擊。

                if not item:
                    x, y, widget = evt.x, evt.y, evt.widget
                    item = widget.identify_row(y)
                    column = widget.identify_column(x)
                if not column or not item in s._items:  # 在工作日行中單擊或僅在列外單擊。
                    return
                item_values = widget.item(item)['values']  # 點選的日期該周
                if not len(item_values):  # 該行是空的。
                    return
                text = item_values[int(column[1]) - 1]  # text 為選擇日期
                if not text:  # 日期為空
                    return
                bbox = widget.bbox(item, column)
                if not bbox:  # 日曆尚未出現
                    self.window.after(20, lambda: s._pressed(item=item, column=column, widget=widget))
                    return
                text = '%02d' % text
                s._selection = (text, item, column)
                s._show_select(text, bbox)

            def _prev_month(s):
                s._canvas.place_forget()
                s._selection = None

                s._date = s._date - timedelta(days=1)
                s._date = datetime(s._date.year, s._date.month, 1)
                s.CB_year.set(s._date.year)
                s.CB_month.set(s._date.month)
                s._update()

            def _next_month(s):
                s._canvas.place_forget()
                s._selection = None

                year, month = s._date.year, s._date.month
                s._date = s._date + timedelta(
                    days=calendar.monthrange(year, month)[1] + 1)
                s._date = datetime(s._date.year, s._date.month, 1)
                s.CB_year.set(s._date.year)
                s.CB_month.set(s._date.month)
                s._update()

            def _update(s, event=None, key=None):
                if key and event.keysym != 'Return': return
                year = int(s.CB_year.get())
                month = int(s.CB_month.get())
                if year == 0 or year > 9999: return
                s._canvas.place_forget()
                s._date = datetime(year, month, 1)
                s._build_calendar()  # 重建日曆

                if year == datetime.now().year and month == datetime.now().month:
                    day = datetime.now().day
                    for _item, day_list in enumerate(s._cal.monthdayscalendar(year, month)):
                        if day in day_list:
                            item = 'I00' + str(_item + 2)
                            column = '#' + str(day_list.index(day) + 1)
                            self.window.after(100, lambda: s._pressed(item=item, column=column, widget=s._calendar))

            def _exit(s, confirm=False):

                global date_list
                date_list = []
                if not confirm:
                    s._selection = None
                year = s._date.year
                month = s._date.month
                self.enydate.insert(1.0, (str(year) + "/" + str(month) + "/" + str(int(s._selection[0])) + "\n"))
                if len(date_list) == 0:
                    date_list.append(self.enydate.get(1.0, "end"))
                else:
                    change_date = date_list[0].replace(" ", "").split("\n")
                    for i in range(len(change_date)):
                        if len(change_date[i]) == 10:
                            if change_date[i][:11] == self.enydate.get(1.0, 2.0):
                                tkmessage.showerror(title="日期重複", message="此日期已選擇")
                            else:
                                date_list.append(self.enydate.get(1.0, "end"))
                # print(date_list)

            def _main_judge(s):
                try:
                    if self.window.focus_displayof() == None or 'toplevel' not in str(self.window.focus_displayof()):
                        s._exit()
                    else:
                        self.window.after(10, s._main_judge)
                except:
                    self.window.after(10, s._main_judge)

            def selection(s):
                if not s._selection:
                    return None
                year = s._date.year
                month = s._date.month
                return str(datetime(year, month, int(s._selection[0])))[:10]

            def Input_judgment(s, content):
                if content.isdigit() or content == "":
                    return True
                else:
                    return False

        cal = Calendar()

    def click_btnYes(self):
        self.window.destroy()

        change_date = date_list[0].replace(" ", "").split("\n")
        new_wb = openpyxl.Workbook()
        ws1 = new_wb.active
        name = meeting_name.get()
        ws1.title = name

        for i in range(len(change_date)):
            ws1.cell(row=1, column=i + 2).value = change_date[i]

        new_wb.save('%s.xlsx' % str(name))

        global meeting_names, finish_meeting
        meeting_names.append(meeting_name.get())
        finish_meeting.append('unfinished')

        for i in range(len(meeting_names)):
            sheet_names.cell(row=i + 1, column=1).value = meeting_names[i]
            sheet_names.cell(row=i + 1, column=2).value = finish_meeting[i]

        names.save('會議名稱.xlsx')

        self.page1.destroy()
        Page1()

    def click_btn_meetings(self, a):
        global name, wb_record, sheet_time, finish_meeting, location
        name = meeting_names[a]
        finish = finish_meeting[a]
        wb_record = openpyxl.load_workbook('%s.xlsx' % name)
        sheet_time = wb_record[name]
        location = a

        if finish == 'finished':
            self.page1.destroy()
            Page9()
        else:
            self.page1.destroy()
            Page3()


class Page3:
    def _from_rgb(self, rgb):
        return "#%02x%02x%02x" % rgb

    def __init__(self, master=None):
        color_1 = self._from_rgb((68, 84, 106))  # 藍黑色
        color_2 = self._from_rgb((208, 224, 227))  # 湖水藍
        color_3 = self._from_rgb((255, 217, 102))  # 淡橘

        self.root = master
        self.page3 = tk.Frame(self.root, width=1000, height=700, bg=color_2)
        self.page3.master.title("Page3")
        self.page3.grid()

        f1 = tkfont.Font(size=30, family="源泉圓體 B")
        f2 = tkfont.Font(size=20, family="源泉圓體 M")
        f3 = tkfont.Font(size=15, family="源泉圓體 M")

        self.lab_title_3 = tk.Label(self.page3, text=' ' + name, height=1, width=15, font=f1, bg=color_1, fg='white',
                                    anchor='w')
        self.btn_createtime = tk.Button(self.page3, text="新增你的時間", height=1, width=18, font=f2, fg=color_1,
                                        relief='solid', command=self.click_btn_createtime)
        self.btn_times = tk.Button(self.page3, text="查看所有人的時間", height=1, width=18, font=f2, fg=color_1, relief='solid',
                                   command=self.click_btn_times)
        self.btn_meetingrecord = tk.Button(self.page3, text="紀錄會議", height=1, width=18, font=f2, fg=color_1,
                                           relief='solid', command=self.click_btn_meetingrecord)
        self.btn_back = tk.Button(self.page3, text="返回", height=1, font=f3, command=self.click_btn_back, bg=color_3)

        self.lab_title_3.place(x=0, y=50)
        self.btn_createtime.place(relx=0.5, y=200, anchor='center')
        self.btn_times.place(relx=0.5, y=300, anchor='center')
        self.btn_meetingrecord.place(relx=0.5, y=400, anchor='center')
        self.btn_back.place(relx=0.5, y=600, anchor='center')

    def click_btn_createtime(self):
        self.page3.destroy()
        Page4()

    def click_btn_times(self):
        self.page3.destroy()
        Page5()

    def click_btn_meetingrecord(self):
        self.page3.destroy()
        Page6()

    def click_btn_back(self):
        self.page3.destroy()
        Page1()


class Page4:
    def _from_rgb(self, rgb):
        return "#%02x%02x%02x" % rgb

    def __init__(self, master=None):
        color_1 = self._from_rgb((68, 84, 106))  # 藍黑色
        color_2 = self._from_rgb((208, 224, 227))  # 湖水藍
        color_3 = self._from_rgb((255, 217, 102))  # 淡橘

        self.root = master
        self.page4 = tk.Frame(self.root, width=1000, height=700, bg=color_2)
        self.page4.master.title("Page4")
        self.page4.grid()

        f1 = tkfont.Font(size=30, family="源泉圓體 B")
        f2 = tkfont.Font(size=20, family="源泉圓體 M")
        f3 = tkfont.Font(size=15, family="源泉圓體 M")

        global var_name, var_selectall

        self.lab_title = tk.Label(self.page4, text=' 新增你的時間', height=1, width=15, font=f1,
                                  bg=color_1, fg='white', anchor='w').place(x=0, y=50)
        self.lab_name = tk.Label(self.page4, text='姓名：', font=f2, bg=color_2)

        var_name = tk.StringVar()
        self.int_name = tk.Entry(self.page4, textvariable=var_name, width=18, font=f2)
        self.btn_yes = tk.Button(self.page4, text='確定', bg=color_3, height=1, font=f3,
                                 command=self.click_btn_yes)
        self.btn_back = tk.Button(self.page4, text='返回', bg=color_3, height=1, font=f3,
                                  command=self.click_btn_back)

        var_selectall = tk.IntVar()
        self.btn_selectall = tk.Checkbutton(self.page4, onvalue=1, offvalue=0, variable=var_selectall, font=f1,
                                            bg=color_2, command=self.click_selectall)
        self.lab_selectall = tk.Label(self.page4, text='全選', font=f3, bg=color_2)

        self.lab_name.place(x=100, y=170)
        self.int_name.place(x=100, y=210, height=30)
        self.btn_back.place(x=440, y=620)
        self.btn_yes.place(x=515, y=620)
        self.btn_selectall.place(x=855, y=35)
        self.lab_selectall.place(x=878, y=50)

        global chk_btns, dates
        chk_btns = []
        dates = []

        for i in range(1, len(list(sheet_time.rows)[0])):
            if str(list(sheet_time.rows)[0][i].value) != 'None':
                dates.append(str(list(sheet_time.rows)[0][i].value).replace("2020/", ""))

        for i in range(len(dates) + 1):
            if i != 0:
                chk_btns.append([])
            for j in range(17):
                if i == 0:
                    if j == 0:
                        tk.Label(self.page4, relief='solid', borderwidth=1, width=10, height=2, bg=color_2).place(x=480, y=80)
                    else:
                        tk.Label(self.page4, text=str(6 + j) + ':00-' + str(7 + j) + ':00', relief='solid',
                                 borderwidth=1, width=10, height=2, bg=color_2).place(x=480, y=80 + 30 * j)
                else:
                    tk.Label(self.page4, relief='solid', borderwidth=1, width=7, height=2,
                             bg=color_2).place(x=501 + 52 * i, y=80 + 30 * j)
                    if j == 0:
                        tk.Label(self.page4, text=dates[i - 1], relief='solid', borderwidth=1, width=7, height=2,
                                 bg=color_2).place(x=501 + 52 * i, y=80)
                    else:
                        var_i = tk.IntVar()
                        tk.Checkbutton(self.page4, onvalue=1, offvalue=0, variable=var_i,
                                       bg=color_2).place(x=518 + 52 * i, y=85 + 30 * j)
                        chk_btns[i - 1].append(var_i)

    def click_btn_yes(self):
        if var_name.get() == "":
            tkmessage.showerror(title="請輸入姓名", message="請輸入姓名！")
        else:
            member = str(sheet_time.cell(row=18, column=1).value)
            if member == 'None':
                list_member = [var_name.get()]
                for i in range(len(dates)):
                    for j in range(16):
                        if chk_btns[i][j].get() == 1:
                            available = str(sheet_time.cell(row=j + 2, column=i + 2).value)
                            if available == 'None':
                                list_available = [var_name.get()]
                            else:
                                list_available = available.split(',')
                                list_available.append(var_name.get())
                            sheet_time.cell(row=j + 2, column=i + 2).value = ",".join(list_available)
            else:
                list_member = member.split(',')
                if var_name.get() not in list_member:
                    list_member.append(var_name.get())
                    for i in range(len(dates)):
                        for j in range(16):
                            if chk_btns[i][j].get() == 1:
                                available = str(sheet_time.cell(row=j + 2, column=i + 2).value)
                                if available == 'None':
                                    list_available = [var_name.get()]
                                else:
                                    list_available = available.split(',')
                                    list_available.append(var_name.get())
                                sheet_time.cell(row=j + 2, column=i + 2).value = ",".join(list_available)
                else:
                    for i in range(len(dates)):
                        for j in range(16):
                            available = str(sheet_time.cell(row=j + 2, column=i + 2).value)
                            if chk_btns[i][j].get() == 1:
                                if available == 'None':
                                    list_available = [var_name.get()]
                                else:
                                    list_available = available.split(',')
                                    if var_name.get() not in available:
                                        list_available.append(var_name.get())
                                sheet_time.cell(row=j + 2, column=i + 2).value = ",".join(list_available)
                            else:
                                if var_name.get() in available:
                                    available = available.replace(var_name.get(), "").replace(",,", ",").strip(",")
                                    sheet_time.cell(row=j + 2, column=i + 2).value = available

            sheet_time.cell(row=18, column=1).value = ",".join(list_member)
            wb_record.save('%s.xlsx' % name)

            self.page4.destroy()
            Page5()

    def click_btn_back(self):
        self.page4.destroy()
        Page3()

    def click_selectall(self):
        if var_selectall.get() == 1:
            for i in range(int(sheet_time.max_column) - 1):
                for j in range(16):
                    chk_btns[i][j].set(1)
        else:
            for i in range(int(sheet_time.max_column) - 1):
                for j in range(16):
                    chk_btns[i][j].set(0)


class Page5:
    def _from_rgb(self, rgb):
        return "#%02x%02x%02x" % rgb

    def __init__(self, master=None):
        color_1 = self._from_rgb((68, 84, 106))  # 藍黑色
        color_2 = self._from_rgb((208, 224, 227))  # 湖水藍
        color_3 = self._from_rgb((255, 217, 102))  # 淡橘

        self.root = master
        self.page5 = tk.Frame(self.root, width=1000, height=700, bg=color_2)
        self.page5.master.title("Page5")
        self.page5.grid()

        f1 = tkfont.Font(size=30, family="源泉圓體 B")
        f2 = tkfont.Font(size=12, family="源泉圓體 M")
        f3 = tkfont.Font(size=15, family="源泉圓體 M")
        f4 = tkfont.Font(size=10, family="源泉圓體 M")

        self.lab_title = tk.Label(self.page5, text=" 查看所有人的時間", height=1, width=15, font=f1, bg=color_1,
                                  fg='white', anchor='w').place(x=0, y=50)

        self.btn_yes = tk.Button(self.page5, text='確定', bg=color_3, height=1, font=f3, command=self.click_btn_yes)
        self.lab_allmembers = tk.Label(self.page5, text='all members', font=f2, bg=color_2)
        self.lab_select = tk.Label(self.page5, text='select', font=f2, bg=color_2)
        self.lab_available = tk.Label(self.page5, text='available', font=f2, bg=color_2)
        self.lab_unavailable = tk.Label(self.page5, text='unavailable', font=f2, bg=color_2)

        self.btn_yes.place(x=475, y=620)
        self.lab_allmembers.place(x=100, y=400)
        self.lab_select.place(x=330, y=400)
        self.lab_available.place(x=320, y=125)
        self.lab_unavailable.place(x=102, y=125)

        self.scroll_available = tk.Scrollbar(self.page5)
        self.scroll_unavailable = tk.Scrollbar(self.page5)
        self.scroll_allmembers = tk.Scrollbar(self.page5)
        self.scroll_color = tk.Scrollbar(self.page5)

        self.scroll_available.place(x=413, y=152, relheight=0.33)
        self.scroll_unavailable.place(x=208, y=152, relheight=0.33)
        self.scroll_allmembers.place(x=208, y=426, relheight=0.194)
        self.scroll_color.place(x=413, y=426, relheight=0.194)

        var_available = tk.StringVar()
        self.lst_available = tk.Listbox(self.page5, listvariable=var_available, font=f2, width=14, height=12,
                                        yscrollcommand=self.scroll_available.set)
        self.lst_available.place(x=285, y=151)

        var_unavailable = tk.StringVar()
        self.lst_unavailable = tk.Listbox(self.page5, listvariable=var_unavailable, font=f2, width=14, height=12,
                                          yscrollcommand=self.scroll_unavailable.set)
        self.lst_unavailable.place(x=80, y=151)

        var_allmembers = tk.StringVar()
        self.lst_allmembers = tk.Listbox(self.page5, listvariable=var_allmembers, font=f2, width=14, height=7,
                                         yscrollcommand=self.scroll_allmembers.set)
        self.lst_allmembers.place(x=80, y=425)

        all_members = str(sheet_time.cell(row=18, column=1).value).split(',')
        how_many_people = len(all_members)

        for member in all_members:
            self.lst_allmembers.insert("end", member)

        self.btn_try = tk.Button(self.page5, text='try', font=f4, width=5, command=self.click_btn_try)
        self.btn_reset = tk.Button(self.page5, text='reset', font=f4, width=5, command=self.click_btn_reset)

        self.btn_try.place(x=298, y=570)
        self.btn_reset.place(x=363, y=570)

        self.lst_color = tk.Listbox(self.page5, width=14, height=7, font=f2, selectmode=tk.MULTIPLE,
                                    yscrollcommand=self.scroll_color.set)
        self.lst_color.place(x=285, y=425)
        self.color = str()

        global color_list, people_list, btn_list, dates
        color_list = []
        people_list = []
        btn_list = []

        dates = []

        for i in range(1, len(list(sheet_time.rows)[0])):
            if str(list(sheet_time.rows)[0][i].value) != 'None':
                dates.append(str(list(sheet_time.rows)[0][i].value).replace("2020/", ""))

        for i in range(how_many_people):
            self.lst_color.insert('end', i + 1)
            self.color = colors[int(30 / int(how_many_people)) * i]
            color_list.append(self.color)
            self.lst_color.itemconfig('end', bg=self.color, selectbackground=self.color)

        for i in range(len(dates) + 1):
            if i != 0:
                people_list.append([])
                btn_list.append([])
            for j in range(17):
                if i == 0:
                    if j == 0:
                        tk.Label(self.page5, relief='solid', borderwidth=1, width=10, height=2, bg=color_2).place(x=480,
                                                                                                                  y=80)
                    else:
                        tk.Label(self.page5, text=str(6 + j) + ':00-' + str(7 + j) + ':00', relief='solid',
                                 borderwidth=1, width=10, height=2, bg=color_2).place(x=480, y=80 + 30 * j)
                else:
                    tk.Label(self.page5, relief='solid', borderwidth=1, width=7, height=2,
                             bg=color_2).place(x=501 +52 * i, y=80 + 30 * j)

                    if j == 0:
                        tk.Label(self.page5, text=dates[i - 1], relief='solid', borderwidth=1, width=7, height=2,
                                 bg=color_2).place(x=501 + 52 * i, y=80)
                    else:
                        how_many_available = len(str(sheet_time.cell(row=j + 1, column=i + 1).value).split(','))
                        if str(sheet_time.cell(row=j + 1, column=i + 1).value) == 'None':
                            how_many_available = 0
                        if how_many_available != 0:
                            self.btn = tk.Button(self.page5, bg=str(color_list[how_many_available - 1]), width=6,
                                                 height=1, command=lambda a=i, b=j: self.click_btn(a, b))
                        else:
                            self.btn = tk.Button(self.page5, bg='white', width=6, height=1,
                                                 command=lambda a=i, b=j: self.click_btn(a, b))
                        self.btn.place(x=502 + 52 * i, y=83 + 30 * j)
                        people_list[i - 1].append(how_many_available)
                        btn_list[i - 1].append(self.btn)

        self.scroll_available.config(command=self.lst_available.yview)
        self.scroll_unavailable.config(command=self.lst_unavailable.yview)
        self.scroll_allmembers.config(command=self.lst_allmembers.yview)
        self.scroll_color.config(command=self.lst_color.yview)

    def click_btn_try(self):
        number_choose = list()
        number_choose_tuple = self.lst_color.curselection()
        for i in number_choose_tuple:
            number_choose.append(self.lst_color.get(i))

        for i in range(len(dates)):
            for j in range(16):
                for k in range(len(number_choose)):
                    if k == 0:
                        if int(people_list[i][j]) == number_choose[k]:
                            btn_list[i][j].config(bg=str(color_list[int(number_choose[k]) - 1]))
                        else:
                            btn_list[i][j].config(bg='white')
                    else:
                        if int(people_list[i][j]) == number_choose[k]:
                            btn_list[i][j].config(bg=str(color_list[int(number_choose[k]) - 1]))

    def click_btn_reset(self):
        number_choose = list()
        number_choose_tuple = self.lst_color.curselection()
        for i in number_choose_tuple:
            number_choose.append(int(self.lst_color.get(i)))

        for k in number_choose:
            self.lst_color.delete(k - 1)
            self.lst_color.insert(k - 1, k)
            self.lst_color.itemconfig(k - 1, bg=color_list[k - 1])

        for i in range(1, len(dates) + 1):
            for j in range(1, 17):
                how_many_available = len(str(sheet_time.cell(row=j + 1, column=i + 1).value).split(','))
                if str(sheet_time.cell(row=j + 1, column=i + 1).value) == 'None':
                    how_many_available = 0
                if how_many_available != 0:
                    btn_list[i - 1][j - 1].config(bg=str(color_list[how_many_available - 1]))
                    people_list[i - 1].append(how_many_available)
                else:
                    btn_list[i - 1][j - 1].config(bg='white')
                    people_list[i - 1].append(how_many_available)

    def click_btn_yes(self):
        self.page5.destroy()
        Page3()

    def click_btn(self, a, b):
        self.lst_available.delete(0, "end")
        self.lst_unavailable.delete(0, "end")

        all_members = str(sheet_time.cell(row=2, column=10).value).split(',')
        available_member = str(sheet_time.cell(row=b + 1, column=a + 1).value).split(',')

        for member in available_member:
            self.lst_available.insert("end", member)

        for member in all_members:
            if member not in available_member:
                self.lst_unavailable.insert("end", member)


class Page6:
    def _from_rgb(self, rgb):
        return "#%02x%02x%02x" % rgb

    def __init__(self, master=None):
        color_1 = self._from_rgb((68, 84, 106))  # 藍黑色
        color_2 = self._from_rgb((208, 224, 227))  # 湖水藍
        color_3 = self._from_rgb((255, 217, 102))  # 淡橘

        self.root = master
        self.page6 = tk.Frame(self.root, width=1000, height=700, bg=color_2)
        self.page6.master.title("Page6")
        self.page6.grid()

        f1 = tkfont.Font(size=30, family="源泉圓體 B")
        f2 = tkfont.Font(size=20, family="源泉圓體 M")
        f3 = tkfont.Font(size=15, family="源泉圓體 M")

        self.lblTitle_6 = tk.Label(self.page6, text=' ' + name + '-紀錄', height=1, width=15, font=f1, bg=color_1,
                                   fg='white', anchor='w').place(x=0, y=50)
        self.btn6_1 = tk.Button(self.page6, text="出缺勤", height=1, width=18, font=f2, relief='solid', fg=color_1,
                                command=self.click_btn6_1)
        self.btn6_2 = tk.Button(self.page6, text="會議記錄", height=1, width=18, font=f2, relief='solid', fg=color_1,
                                command=self.click_btn6_2)
        self.btn6_3 = tk.Button(self.page6, text="結束會議", height=1, width=18, font=f2, relief='solid', fg=color_1,
                                command=self.click_btn6_3)
        self.btn6_4 = tk.Button(self.page6, text="返回", font=f3, bg=color_3, command=self.click_btn6_4)

        self.btn6_1.place(relx=0.5, y=200, anchor='center')
        self.btn6_2.place(relx=0.5, y=300, anchor='center')
        self.btn6_3.place(relx=0.5, y=400, anchor='center')
        self.btn6_4.place(relx=0.5, y=600, anchor='center')

    def click_btn6_1(self):
        self.page6.destroy()
        Page7()

    def click_btn6_2(self):
        self.page6.destroy()
        Page8()

    def click_btn6_3(self):
        message = tkmessage.askokcancel(title="確定結束會議？", message="結束會議後，您將無法作任何更動")
        if message:
            global absence, mission
            absence = []
            mission = []
            sheet_names.cell(column=2, row=location + 1).value = 'finished'
            names.save('會議名稱.xlsx')
            self.page6.destroy()
            Page9()

    def click_btn6_4(self):
        self.page6.destroy()
        Page3()


class Page7:
    def _from_rgb(self, rgb):
        return "#%02x%02x%02x" % rgb

    def __init__(self, master=None):
        color_1 = self._from_rgb((68, 84, 106))  # 藍黑色
        color_2 = self._from_rgb((208, 224, 227))  # 湖水藍
        color_3 = self._from_rgb((255, 217, 102))  # 淡橘

        self.root = master
        self.page7 = tk.Frame(self.root, width=1000, height=700, bg=color_2)
        self.page7.master.title("Page8")
        self.page7.grid()
        # 因為scrollbar不能直接應用在frame上，所以就創了一個canvas，在canvas上加scrollbar，然後再把canvas放進frame
        self.canvas = tk.Canvas(self.page7, width=1000, height=700, bg=color_2)
        self.canvas.grid()
        self.slb = tk.Scrollbar(self.page7, orient='vertical')
        self.slb.place(x=980, width=20, height=700)
        self.canvas.configure(yscrollcommand=self.slb.set)
        self.slb.configure(command=self.canvas.yview)
        self.frame_context = tk.Frame(self.canvas, width=1000, height=700, bg=color_2)
        self.canvas.create_window((0, 0), window=self.frame_context, anchor='nw')

        f1 = tkfont.Font(size=30, family="源泉圓體 B")
        f2 = tkfont.Font(size=12, family="源泉圓體 M")
        f3 = tkfont.Font(size=15, family="源泉圓體 M")

        self.lab_title = tk.Label(self.page7, width=15, height=1, text=' 出缺勤', font=f1, fg='white', bg=color_1,
                                  anchor='w').place(x=0, y=50)

        self.btn7_1 = tk.Button(self.page7, text="確定", font=f3, bg=color_3, command=self.click_btn7_1)
        self.btn7_1.place(x=850, y=60, anchor='center')

        self.lab7_1 = tk.Label(self.page7, text="準時", font=f2, bg=color_2).place(x=179, y=125)
        self.lab7_2 = tk.Label(self.page7, text="遲到", font=f2, bg=color_2).place(x=279, y=125)
        self.lab7_3 = tk.Label(self.page7, text="未出席", font=f2, bg=color_2).place(x=372, y=125)

        self.lab7_4 = tk.Label(self.page7, text="是", font=f2, bg=color_2).place(x=700, y=125)
        self.lab7_5 = tk.Label(self.page7, text="否", font=f2, bg=color_2).place(x=770, y=125)
        self.lab7_6 = tk.Label(self.page7, text="無任務", font=f2, bg=color_2).place(x=823, y=125)

        try:
            sheet = wb_record['出缺勤']
        except KeyError:
            sheet = wb_record.create_sheet('出缺勤')

        global absence, mission, member_list

        member_list = str(sheet_time.cell(row=18, column=1).value).split(',')
        canvas_height = 160  # 計算最後頁面會有多長
        for i in range(len(member_list)):
            sheet.cell(row=i + 1, column=1).value = member_list[i]
            tk.Label(self.page7, text=member_list[i], font=f2, bg=color_2).place(x=115, y=168 + 35 * i, anchor='center')
            tk.Label(self.page7, text='是否完成指派任務？', font=f2, bg=color_2).place(x=515, y=155 + 35 * i)

            if len(absence) != len(member_list):
                var_absence = tk.IntVar()
                absence.append(var_absence)
            else:
                var_absence = absence[i]
            tk.Radiobutton(self.page7, variable=var_absence, value=1, bg=color_2).place(x=185, y=155 + 35 * i)
            tk.Radiobutton(self.page7, variable=var_absence, value=2, bg=color_2).place(x=285, y=155 + 35 * i)
            tk.Radiobutton(self.page7, variable=var_absence, value=3, bg=color_2).place(x=385, y=155 + 35 * i)

            if len(mission) != len(member_list):
                var_mission = tk.IntVar()
                mission.append(var_mission)
            else:
                var_mission = mission[i]
            tk.Radiobutton(self.page7, variable=var_mission, value=1, bg=color_2).place(x=700, y=155 + 35 * i)
            tk.Radiobutton(self.page7, variable=var_mission, value=2, bg=color_2).place(x=770, y=155 + 35 * i)
            tk.Radiobutton(self.page7, variable=var_mission, value=3, bg=color_2).place(x=840, y=155 + 35 * i)
            canvas_height += 35

        wb_record.save('%s.xlsx' % name)
        if canvas_height > 700:
            self.canvas.configure(scrollregion=(0,0,1000,canvas_height))
        else:
            self.canvas.configure(scrollregion=(0,0,1000,700))

    def click_btn7_1(self):
        sheet = wb_record['出缺勤']

        for i in range(len(member_list)):
            if absence[i].get() == 1:
                sheet.cell(row=i + 1, column=2).value = "準時"
            if absence[i].get() == 2:
                sheet.cell(row=i + 1, column=2).value = "遲到"
            if absence[i].get() == 3:
                sheet.cell(row=i + 1, column=2).value = "未出席"
            if mission[i].get() == 1:
                sheet.cell(row=i + 1, column=3).value = "完成任務"
            if mission[i].get() == 2:
                sheet.cell(row=i + 1, column=3).value = "未完成任務"
            if mission[i].get() == 3:
                sheet.cell(row=i + 1, column=3).value = "無任務"

        wb_record.save('%s.xlsx' % name)

        self.page7.destroy()
        Page6()


class Page8:
    def _from_rgb(self, rgb):
        return "#%02x%02x%02x" % rgb

    def __init__(self, master=None):
        color_1 = self._from_rgb((68, 84, 106))  # 藍黑色
        color_2 = self._from_rgb((208, 224, 227))  # 湖水藍
        color_3 = self._from_rgb((255, 217, 102))  # 淡橘

        self.root = master
        self.page8 = tk.Frame(self.root, width=1000, height=700, bg=color_2)
        self.page8.master.title("Page8")
        self.page8.grid()

        f1 = tkfont.Font(size=30, family="源泉圓體 B")
        f2 = tkfont.Font(size=12, family="源泉圓體 M")
        f3 = tkfont.Font(size=15, family="源泉圓體 M")

        self.lbl8_1 = tk.Label(self.page8, text=" 會議記錄", height=1, width=15, font=f1, fg='white', bg=color_1,
                               anchor='w')
        self.record = tk.Text(self.page8, height=23, width=90, font=f2)
        self.btn8 = tk.Button(self.page8, text='確認', font=f3, bg=color_3, command=self.click_btn8)
        self.lbl8_2 = tk.Label(self.page8, text="會議時間：", font=f3, bg=color_2)

        global var_times
        var_times = tk.StringVar()
        self.entry8 = tk.Entry(self.page8, textvariable=var_times, width=18, font=f3)

        global meeting_record
        try:
            meeting_record = wb_record['Meeting record']
        except KeyError:
            meeting_record = wb_record.create_sheet('Meeting record')

        if str(meeting_record.cell(row=1, column=1).value) != 'None':
            self.record.insert("1.0", str(meeting_record.cell(row=1, column=1).value))

        self.lbl8_1.place(x=0, y=50)
        self.record.place(relx=0.5, rely=0.5, anchor='center')
        self.btn8.place(relx=0.5, y=620, anchor='center')
        self.entry8.place(x=705, y=105)
        self.lbl8_2.place(x=600, y=105)

    def click_btn8(self):
        meeting_record.cell(row=1, column=1).value = self.record.get("1.0", "end")
        meeting_record.cell(row=1, column=2).value = var_times.get()
        wb_record.save('%s.xlsx' % name)
        self.page8.destroy()
        Page6()


class Page9:
    def _from_rgb(self, rgb):
        return "#%02x%02x%02x" % rgb

    def __init__(self, master=None):
        color_1 = self._from_rgb((68, 84, 106))  # 藍黑色
        color_2 = self._from_rgb((208, 224, 227))  # 湖水藍
        color_3 = self._from_rgb((255, 217, 102))  # 淡橘

        self.root = master
        self.page9 = tk.Frame(self.root, width=1000, height=700, bg=color_2)
        self.page9.master.title("Page9")
        self.page9.grid()

        f1 = tkfont.Font(size=30, family="源泉圓體 B")
        f2 = tkfont.Font(size=12, family="源泉圓體 M")
        f3 = tkfont.Font(size=15, family="源泉圓體 M")

        ws_1 = wb_record['Meeting record']
        ws_2 = wb_record['出缺勤']

        self.lbl_title9 = tk.Label(self.page9, text=' 結束會議', height=1, width=10, font=f1, fg='white', bg=color_1,
                                   anchor='w').place(x=0, y=50)
        self.lbl9_1 = tk.Label(self.page9, text="會議名稱：", font=f2, bg=color_2)
        self.lbl9_2 = tk.Label(self.page9, text=name, font=f2, bg=color_2)
        self.lbl9_3 = tk.Label(self.page9, text="會議時間：", font=f2, bg=color_2)
        self.lbl9_4 = tk.Label(self.page9, text=str(ws_1.cell(row=1, column=2).value), font=f2, bg=color_2)
        self.lbl9_5 = tk.Label(self.page9, text="會議記錄：", font=f2, bg=color_2)
        self.lbl9_6 = tk.Label(self.page9, text=str(ws_1.cell(row=1, column=1).value), font=f2, bg=color_2)
        self.lbl9_7 = tk.Label(self.page9, text="成員名單：", font=f2, bg=color_2)
        self.lbl9_9 = tk.Label(self.page9, text="出缺勤：", font=f2, bg=color_2)
        self.lbl9_11 = tk.Label(self.page9, text="是否完成指派任務？", font=f2, bg=color_2)
        self.btn9 = tk.Button(self.page9, text="確定", font=f3, bg=color_3, command=self.click_btn9_1)

        self.lbl9_1.place(x=100, y=130)
        self.lbl9_2.place(x=180, y=130)
        self.lbl9_3.place(x=100, y=180)
        self.lbl9_4.place(x=180, y=180)
        self.lbl9_5.place(x=100, y=230)
        self.lbl9_6.place(x=180, y=230)
        self.lbl9_7.place(x=100, y=430)
        self.lbl9_9.place(x=100, y=480)
        self.lbl9_11.place(x=100, y=530)
        self.btn9.place(relx=0.5, y=620, anchor='center')

        times = wb_record[name]
        member_list = str(times.cell(row=18, column=1).value).split(',')
        for i in range(len(member_list)):
            self.lbl9_8 = tk.Label(self.page9, text=str(ws_2.cell(row=i + 1, column=1).value), font=f2,
                                   bg=color_2).place(x=280 + 70 * i, y=430)
            self.lbl9_10 = tk.Label(self.page9, text=str(ws_2.cell(row=i + 1, column=2).value), font=f2,
                                    bg=color_2).place(x=280 + 70 * i, y=480)
            self.lbl9_12 = tk.Label(self.page9, text=str(ws_2.cell(row=i + 1, column=3).value), font=f2,
                                    bg=color_2).place(x=280 + 70 * i, y=530)

    def click_btn9_1(self):
        self.page9.destroy()
        Page1()


root.geometry("1000x700")
root.resizable(0, 0)
Page0(root)
root.mainloop()
